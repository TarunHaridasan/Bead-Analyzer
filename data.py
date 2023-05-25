import cv2
import os
from math import sqrt
from openpyxl import Workbook
from openpyxl.styles import Font

#Find the center point of a rectangle
def findCntrPnt(rect):
    x1,y1,w,h = rect[0], rect[1], rect[2], rect[3]
    x2 = x1 + w
    y2 = y1 + h
    xa = (x1+x2)/2
    ya = (y1+y2)/2
    return [xa, ya]

#Calculate the distance between 2 points
def dist(start, end):
    x1,y1 = start[0], start[1]
    x2,y2 = end[0], end[1]
    return sqrt((x2-x1)**2 + (y2-y1)**2)

#Main data wrapper
class Data:
    #Load all every 5th image in the folder
    def __init__(self, fp):
        self.fp = fp
        self.images = []
        self.boundingBoxes = []  
        files = os.listdir(self.fp)
        for i in range(0,len(files),5):
            item = files[i]
            fp = f'{self.fp}//{item}'
            image = cv2.imread(fp)
            self.images.append(image)
        self.size = len(self.images)
    #Get an image
    def get(self, index):
        return self.images[index]
    #Print all bounding boxes on image
    def print(self, image, highlight = None):
        frame = image.copy()
        for i in range(len(self.boundingBoxes)):
            box = self.boundingBoxes[i]
            x,y,w,h = box[0], box[1], box[2], box[3]
            if i == highlight:
                cv2.rectangle(frame,(x,y),(x+w,y+h),(0, 255, 21),2)
            else:
                cv2.rectangle(frame,(x,y),(x+w,y+h),(255, 0, 21),2)    
        cv2.imshow("Beads", frame)      
        return frame 
    #Initially detect fluorescent beads off an image
    def findBeads(self):
        image = self.get(0).copy()
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blob = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 3, 5)
        #Opening (remove noise)
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5))
        blob = cv2.morphologyEx(blob, cv2.MORPH_OPEN, kernel)
        #Closing
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5))
        blob = cv2.morphologyEx(blob, cv2.MORPH_CLOSE, kernel)
        #Find contours
        blob = 255 - blob
        contours, hier = cv2.findContours(blob, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        #Find 50th percentile area
        areas = []
        for cnt in contours:
            areas.append(cv2.contourArea(cnt))
        areas.sort()
        a = areas[int(len(areas)*0.50)]
        #Remove bounding boxes with area too small
        boundingBoxes = []
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area>a:
                x,y,w,h = cv2.boundingRect(cnt)
                boundingBoxes.append([x,y,w,h])
        self.boundingBoxes = boundingBoxes
        return boundingBoxes
    #Return the bounding box the x,y coordinates are inside
    def findBox(self,x,y):
        for box in self.boundingBoxes:
            x1,y1,w,h = box[0], box[1], box[2], box[3]
            x2,y2 = x1+w, y1+h
            if x>x1 and x<x2 and y>y1 and y<y2: #Inside
                return [x1, y1, w, h]
        return 0
    #Remove bounding box
    def removeBounding(self, event, x, y, flags, params):
        if event == cv2.EVENT_LBUTTONUP:
            box = self.findBox(x,y)
            if box!=0:
                self.boundingBoxes.remove(box)
                self.print(self.get(0))
    #Add bounding box
    def addBounding(self):
        while True:
            x,y,w,h = cv2.selectROI("Beads", self.print(self.get(0)))
            if x+y+w+h == 0:
                self.print(self.get(0))
                return
            self.boundingBoxes.append([x,y,w,h])

#Tracker class for individual beads
class SubTracker():
    def __init__(self, image, bounding): #Initialize tracker
        self.start = bounding
        self.prev = self.start
        self.current = self.start
        self.distance = 0
        self.displacement = 0
        self.frames = 0
        self.tracker = cv2.TrackerCSRT_create()
        self.tracker.init(image, bounding)
        self.active = 1
    def update(self, image): #Update tracker frame
        #print(self.active, self.prev, self.current)
        if not self.active: 
            return
        retval, box = self.tracker.update(image)
        self.prev = self.current
        if retval:            
            self.current = box
            if self.current[0] < 50 or self.current[0]>850 or self.current[1]<50 or self.current[1]>750:
                self.active = 0
            self.updateDistance()
            self.updateFrames()
    def draw(self, image): #Draw bounding boxes on rectangle
        cv2.rectangle(image, self.current, (0, 255, 21), 2)
        cv2.imshow('Bead Analyzer', image)
    def updateDistance(self): #Update the distance
        p1, p2 = findCntrPnt(self.prev), findCntrPnt(self.current)
        x1, y1, x2, y2 = p1[0], p1[1], p2[0], p2[1] 
        distance = sqrt((x2-x1)**2 + (y2-y1)**2)
        self.distance+=distance
    def updateDisplacement(self): #Update the displacement
        p1, p2 = findCntrPnt(self.start), findCntrPnt(self.current)
        x1, y1, x2, y2 = p1[0], p1[1], p2[0], p2[1] 
        displacement = sqrt((x2-x1)**2 + (y2-y1)**2)
        self.displacement=displacement
    def updateFrames(self):
        self.frames+=1
        
    
#Main Wrapper Tracker Clas
class Tracker:
    def __init__(self, image):
        self.trackers = []
        self.image = image
        self.size = 0
    def add(self, box):
        tracker = SubTracker(self.image, box)
        self.trackers.append(tracker)
        self.size=+1
    def remove(self, tracker):
        self.trackers.remove(tracker)
    def update(self, image):
        self.image = image
        for tracker in self.trackers:
            tracker.update(image)
    def updateDistances(self):
        for tracker in self.trackers:
            tracker.updateDistance()
    def updateDisplacements(self):
        for tracker in self.trackers:
            tracker.updateDisplacement()
    def draw(self):
        for tracker in self.trackers:
            tracker.draw(self.image)
    def saveData(self):
        conversion = 2
        fps = 64.47
        wb = Workbook()
        ws = wb.active
        ws.append(["Blob", "Distance (um)", "Displacement (um)", "Frames", "Speed (um/s)", "Velocity (um/s)", "Start Coordinates", "End Coordinates"])
        ws["A1"].font, ws["B1"].font, ws["C1"].font, ws["D1"].font, ws["E1"].font, ws["F1"].font = Font(bold=True), Font(bold=True), Font(bold=True), Font(bold=True), Font(bold=True), Font(bold=True)
        for i in range(len(self.trackers)):
            blob = f'Blob {i}' 
            tracker = self.trackers[i]           
            distance = tracker.distance * conversion
            displacement = tracker.displacement * conversion
            frames = tracker.frames
            speed = distance / (frames*5/fps)
            velocity = displacement / (frames*5/fps)
            initial = tracker.start
            final = tracker.current
            ws.append([blob, distance, displacement, frames, speed, velocity, str(initial), str(final)])
        wb.save("data.xlsx")

        