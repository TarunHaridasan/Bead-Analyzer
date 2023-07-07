from openpyxl import Workbook, load_workbook
import os

def process(ws, dir):
    ws.append(["Bead", "Distance (um)", "Displacement (um)", "Speed (um/s)", "Velocity {um/s)}"])
    for filename in os.listdir(dir):
        fp = f'{dir}\\{filename}'
        print(fp)
        fileWB = load_workbook(fp)
        fileWS = fileWB.active    
        for row in fileWS.iter_rows(min_row=4, max_col=6,values_only=True):
            ws.append(row)

#Init
id = input("Enter the ID: ")
controlFP = input("Enter the Control filepath: ")
nicFP = input("Enter the Nicotine filepath: ")
pgvgFP = input("Enter the PG-VG filepath: ")

#New Workbook and sheets
wb = Workbook()
ws = wb.active
ws.title = "Control"  
wb.create_sheet("Nicotine")
wb.create_sheet("PG-VG")

#Process
process(wb["Control"], controlFP)
process(wb["Nicotine"], nicFP)
process(wb["PG-VG"], pgvgFP)

#Save
wb.save(f'.//Output//{id}.xlsx')
print("Completed")